#!/usr/bin/env python3
"""90日間アクションプランのXLSXスプレッドシートを生成するスクリプト。

Usage:
    echo '{"title": "...", "phases": [...]}' | python generate_xlsx.py output.xlsx
    python generate_xlsx.py output.xlsx input.json

入力JSONが省されない場合はテンプレートのみ生成する。
"""
import sys
import json
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl が必要です。pip install openpyxl を実行してください。", file=sys.stderr)
    sys.exit(1)


def sanitize_cell(value):
    """Excel formula injection を防止するサニタイズ関数。
    入力JSONは信頼できないものとして扱い、数式トリガー文字を無害化する。"""
    if not isinstance(value, str):
        return value
    if value and value[0] in ('=', '+', '-', '@', '\t', '\r', '\n'):
        return "'" + value
    return value

# Color palette
COLORS = {
    "header": "1F3864",
    "phase1": "D6E4F0",
    "phase2": "E2EFDA",
    "phase3": "FFF2CC",
    "phase4": "FCE4D6",
    "section": "2E75B6",
    "subsection": "BDD7EE",
    "white": "FFFFFF",
    "grey": "F2F2F2",
}
PHASE_COLORS = [COLORS["phase1"], COLORS["phase2"], COLORS["phase3"], COLORS["phase4"]]

THIN_BORDER = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7'),
)

def style_header(ws, row, cols, bg=COLORS["header"]):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name='Arial', bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_row(ws, row, cols, bg=COLORS["white"], bold=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name='Arial', bold=bold, size=10)
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def section_row(ws, row, cols, text, bg=COLORS["section"]):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name='Arial', bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=bg)

def create_actionplan_sheet(wb, data):
    ws = wb.active
    ws.title = "90日アクションプラン"
    COLS = 8
    widths = {'A': 8, 'B': 8, 'C': 14, 'D': 40, 'E': 12, 'F': 10, 'G': 28, 'H': 8}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:H1')
    ws['A1'].value = sanitize_cell(data.get("title", "90日間アクションプラン"))
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color="1F3864")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:H2')
    ws['A2'].value = sanitize_cell(data.get("subtitle", ""))
    ws['A2'].font = Font(name='Arial', size=10, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    r = 4
    headers = ['Day', 'Week', 'Phase', 'タスク', 'カテゴリ', '時間(h)', 'アウトプット', '完了']
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, COLS)
    r += 1

    for pi, phase in enumerate(data.get("phases", [])):
        phase_name = sanitize_cell(phase.get("name", f"Phase {pi+1}"))
        phase_label = sanitize_cell(phase.get("label", ""))
        section_row(ws, r, COLS, f"■ {phase_name}: {phase_label}")
        r += 1
        bg = PHASE_COLORS[pi % len(PHASE_COLORS)]
        for task in phase.get("tasks", []):
            ws.cell(row=r, column=1, value=sanitize_cell(str(task.get("day", ""))))
            ws.cell(row=r, column=2, value=sanitize_cell(task.get("week", "")))
            ws.cell(row=r, column=3, value=sanitize_cell(phase_name))
            ws.cell(row=r, column=4, value=sanitize_cell(task.get("task", "")))
            ws.cell(row=r, column=5, value=sanitize_cell(task.get("category", "")))
            ws.cell(row=r, column=6, value=task.get("hours", ""))
            ws.cell(row=r, column=7, value=sanitize_cell(task.get("output", "")))
            ws.cell(row=r, column=8, value="☐")
            style_row(ws, r, COLS, bg=bg)
            for c in [1, 2, 6, 8]:
                ws.cell(row=r, column=c).alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[r].height = 30
            r += 1

def create_kpi_sheet(wb, data):
    ws = wb.create_sheet("KPI・目標")
    COLS = 6
    widths = {'A': 14, 'B': 30, 'C': 16, 'D': 16, 'E': 16, 'F': 8}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:F1')
    ws['A1'].value = "KPI・目標管理シート"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color="1F3864")
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    r = 3
    for i, h in enumerate(['Phase', 'KPI項目', '目標値', '実績値', '達成率', '完了'], 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, COLS)
    r += 1

    for pi, phase in enumerate(data.get("phases", [])):
        phase_name = sanitize_cell(phase.get("name", f"Phase {pi+1}"))
        bg = PHASE_COLORS[pi % len(PHASE_COLORS)]
        kpis = phase.get("kpis", [])
        start_r = r
        for ki, kpi in enumerate(kpis):
            ws.cell(row=r, column=1, value=phase_name if ki == 0 else "")
            ws.cell(row=r, column=2, value=sanitize_cell(kpi.get("item", "")))
            ws.cell(row=r, column=3, value=sanitize_cell(kpi.get("target", "")))
            ws.cell(row=r, column=4, value="")
            ws.cell(row=r, column=5, value="")
            ws.cell(row=r, column=6, value="☐")
            style_row(ws, r, COLS, bg=bg)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=r, column=6).alignment = Alignment(horizontal='center', vertical='center')
            r += 1
        if len(kpis) > 1:
            ws.merge_cells(start_row=start_r, start_column=1, end_row=r - 1, end_column=1)

def create_tools_sheet(wb, data):
    ws = wb.create_sheet("ツール・費用")
    COLS = 5
    widths = {'A': 24, 'B': 30, 'C': 14, 'D': 14, 'E': 24}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:E1')
    ws['A1'].value = "必要ツール・プラットフォーム一覧"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color="1F3864")
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    r = 3
    for i, h in enumerate(['ツール', '用途', '費用', '必要度', '備考'], 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, COLS)
    r += 1

    for group in data.get("tool_groups", []):
        section_row(ws, r, COLS, sanitize_cell(group.get("label", "")))
        r += 1
        for tool in group.get("tools", []):
            ws.cell(row=r, column=1, value=sanitize_cell(tool.get("name", "")))
            ws.cell(row=r, column=2, value=sanitize_cell(tool.get("purpose", "")))
            ws.cell(row=r, column=3, value=sanitize_cell(tool.get("cost", "")))
            ws.cell(row=r, column=4, value=sanitize_cell(tool.get("priority", "")))
            ws.cell(row=r, column=5, value=sanitize_cell(tool.get("note", "")))
            style_row(ws, r, COLS)
            ws.cell(row=r, column=3).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=r, column=4).alignment = Alignment(horizontal='center', vertical='center')
            cost = tool.get("cost", "")
            if "0円" in str(cost):
                ws.cell(row=r, column=3).font = Font(name='Arial', size=10, color='007A33', bold=True)
            r += 1

def create_legal_sheet(wb, data):
    ws = wb.create_sheet("法務・税務")
    COLS = 4
    widths = {'A': 8, 'B': 35, 'C': 50, 'D': 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:D1')
    ws['A1'].value = "法務・税務チェックリスト"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color="1F3864")
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    r = 3
    for i, h in enumerate(['完了', '項目', '詳細', 'タイミング'], 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, COLS)
    r += 1

    for item in data.get("legal_items", []):
        ws.cell(row=r, column=1, value="☐")
        ws.cell(row=r, column=2, value=sanitize_cell(item.get("item", "")))
        ws.cell(row=r, column=3, value=sanitize_cell(item.get("detail", "")))
        ws.cell(row=r, column=4, value=sanitize_cell(item.get("timing", "")))
        style_row(ws, r, COLS)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 35
        r += 1

def create_review_sheet(wb, weeks=12, kpi_columns=None):
    """週次振り返りシートを生成する。
    kpi_columns: KPI列名のリスト。未指定時は汎用デフォルトを使用。"""
    if kpi_columns is None:
        kpi_columns = ['主要KPI 1', '主要KPI 2', '主要KPI 3']
    headers = ['Week'] + [sanitize_cell(c) for c in kpi_columns] + ['今週の学び', '来週やること']
    ws = wb.create_sheet("週次振り返り")
    COLS = len(headers)
    # 列幅: Week=10, KPI列=14, 学び/やること=30
    col_letters = [chr(ord('A') + i) for i in range(COLS)]
    for i, col in enumerate(col_letters):
        if i == 0:
            ws.column_dimensions[col].width = 10
        elif i >= COLS - 2:
            ws.column_dimensions[col].width = 30
        else:
            ws.column_dimensions[col].width = 14

    end_col = col_letters[-1]
    ws.merge_cells(f'A1:{end_col}1')
    ws['A1'].value = "週次振り返りシート（毎週日曜15分）"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color="1F3864")
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    r = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, COLS)
    r += 1

    phase_week_ranges = [(1, 2), (3, 4), (5, 8), (9, 12)]
    for w in range(1, weeks + 1):
        pi = next((i for i, (s, e) in enumerate(phase_week_ranges) if s <= w <= e), len(phase_week_ranges) - 1)
        bg = PHASE_COLORS[pi % len(PHASE_COLORS)]
        ws.cell(row=r, column=1, value=f"Week {w}")
        style_row(ws, r, COLS, bg=bg)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 35
        r += 1

def generate(output_path, data):
    wb = Workbook()
    create_actionplan_sheet(wb, data)
    create_kpi_sheet(wb, data)
    create_tools_sheet(wb, data)
    create_legal_sheet(wb, data)
    create_review_sheet(wb, data.get("weeks", 12), data.get("review_columns"))
    wb.save(output_path)
    return output_path

def default_template():
    """入力JSONがない場合に使用するデフォルトテンプレート"""
    return {
        "title": "90日間アクションプラン",
        "subtitle": "",
        "weeks": 12,
        "phases": [
            {"name": "Phase 1", "label": "準備期間（Day 1-14）", "tasks": [], "kpis": []},
            {"name": "Phase 2", "label": "MVP・初期テスト（Day 15-28）", "tasks": [], "kpis": []},
            {"name": "Phase 3", "label": "初収益（Day 29-56）", "tasks": [], "kpis": []},
            {"name": "Phase 4", "label": "改善・拡大（Day 57-90）", "tasks": [], "kpis": []},
        ],
        "tool_groups": [],
        "legal_items": [],
    }

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python generate_xlsx.py <output.xlsx> [input.json]", file=sys.stderr)
        sys.exit(1)

    output_path = sys.argv[1]

    if len(sys.argv) >= 3:
        with open(sys.argv[2], 'r', encoding='utf-8') as f:
            data = json.load(f)
    elif not sys.stdin.isatty():
        data = json.load(sys.stdin)
    else:
        data = default_template()

    result = generate(output_path, data)
    print(f"Generated: {result}")
